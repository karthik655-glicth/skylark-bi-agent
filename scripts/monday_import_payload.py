import datetime
import json
import sys

import openpyxl


def clean(value):
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%Y-%m-%d")
    return value


def nonempty(value):
    return value not in (None, "")


def valid_for_kind(value, kind):
    if kind == "date":
        return isinstance(value, (datetime.date, datetime.datetime)) or (isinstance(value, str) and len(value) == 10 and value[4] == "-" and value[7] == "-")
    if kind == "number":
        try:
            float(value)
            return True
        except (TypeError, ValueError):
            return False
    return True


def deal_rows(path, start, end):
    columns = {
        1: ("text_mm6jm53r", "text"), 2: ("text_mm6j33th", "text"), 3: ("color_mm6jhdyv", "status"),
        4: ("date_mm6j8qsn", "date"), 5: ("color_mm6jtc6v", "status"), 6: ("numeric_mm6j7tea", "number"),
        7: ("date_mm6j920e", "date"), 8: ("color_mm6j49a0", "status"), 9: ("text_mm6jf0w4", "text"),
        10: ("dropdown_mm6jy9fn", "dropdown"), 11: ("date_mm6jkm28", "date"),
    }
    rows = list(openpyxl.load_workbook(path, data_only=True).active.iter_rows(min_row=2, values_only=True))[start:end]
    output = []
    for row in rows:
        values = {}
        for index, (column_id, kind) in columns.items():
            value = clean(row[index])
            if not nonempty(value) or not valid_for_kind(value, kind):
                continue
            values[column_id] = {"label": str(value)} if kind == "status" else {"date": str(value)} if kind == "date" else {"labels": [str(value)]} if kind == "dropdown" else str(value)
        output.append({"name": str(row[0] or "Untitled deal")[:255], "columnValues": json.dumps(values), "createLabelsIfMissing": True})
    return output


def work_order_rows(path, start, end):
    columns = {
        1: ("text_mm6jz90g", "text"), 2: ("text_mm6jp0bq", "text"), 3: ("dropdown_mm6jg2ar", "dropdown"),
        5: ("color_mm6js0w4", "status"), 6: ("date_mm6jwa65", "date"), 7: ("date_mm6jxrkj", "date"),
        8: ("dropdown_mm6jfb6k", "dropdown"), 9: ("date_mm6jeh6w", "date"), 10: ("date_mm6jn9vq", "date"),
        11: ("text_mm6j6bwr", "text"), 12: ("dropdown_mm6je8yw", "dropdown"), 13: ("text_mm6j2y3b", "text"),
        17: ("numeric_mm6j6vhr", "number"), 19: ("numeric_mm6jvh4f", "number"), 21: ("numeric_mm6jcj0g", "number"),
        22: ("numeric_mm6j333y", "number"), 24: ("numeric_mm6jeeq7", "number"), 30: ("color_mm6jv63v", "status"),
        35: ("color_mm6jhe89", "status"), 37: ("color_mm6jtfsr", "status"),
    }
    worksheet = openpyxl.load_workbook(path, data_only=True).active
    headers = [str(value or f"column_{index}") for index, value in enumerate(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))[start:end]
    output = []
    for row in rows:
        values = {}
        for index, (column_id, kind) in columns.items():
            value = clean(row[index])
            if not nonempty(value) or not valid_for_kind(value, kind):
                continue
            values[column_id] = {"label": str(value)} if kind == "status" else {"date": str(value)} if kind == "date" else {"labels": [str(value)]} if kind == "dropdown" else str(value)
        raw = {headers[index]: clean(value) for index, value in enumerate(row) if nonempty(value)}
        values["long_text_mm6jkb9w"] = {"text": json.dumps(raw, ensure_ascii=False)}
        output.append({"name": str(row[0] or "Untitled work order")[:255], "columnValues": json.dumps(values), "createLabelsIfMissing": True})
    return output


if __name__ == "__main__":
    kind = sys.argv[4] if len(sys.argv) > 4 else "deals"
    generator = work_order_rows if kind == "work_orders" else deal_rows
    print(json.dumps(generator(sys.argv[1], int(sys.argv[2]), int(sys.argv[3]))))
